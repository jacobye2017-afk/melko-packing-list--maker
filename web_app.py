"""
MELKO Packing List Maker - Web Version (Flask)
"""

import os
import io
import zipfile
import tempfile
import shutil
from flask import Flask, request, send_file, render_template_string, jsonify
import converter

app = Flask(__name__)
app.config['MAX_CONTENT_LENGTH'] = 50 * 1024 * 1024  # 50MB max

HTML = r"""
<!DOCTYPE html>
<html class="light" lang="en"><head>
<meta charset="utf-8"/>
<meta content="width=device-width, initial-scale=1.0" name="viewport"/>
<title>MELKO Packing List Maker</title>
<link href="https://fonts.googleapis.com/css2?family=Inter:wght@300;400;500;600;700;800;900&display=swap" rel="stylesheet"/>
<link href="https://fonts.googleapis.com/css2?family=Material+Symbols+Outlined:wght,FILL@100..700,0..1&display=swap" rel="stylesheet"/>
<script src="https://cdn.tailwindcss.com?plugins=forms,container-queries"></script>
<script>
tailwind.config={darkMode:"class",theme:{extend:{colors:{"primary":"#1039b9","primary-container":"#3454d1","on-primary":"#ffffff","on-primary-container":"#d4d9ff","secondary":"#4e6073","secondary-container":"#cfe2f9","on-secondary":"#ffffff","tertiary":"#684000","tertiary-container":"#895500","on-tertiary-container":"#ffd4a5","background":"#f7f9ff","surface":"#f7f9ff","surface-container":"#e3efff","surface-container-low":"#edf4ff","surface-container-lowest":"#ffffff","surface-container-high":"#d9eaff","on-surface":"#091d2e","on-surface-variant":"#444654","outline":"#747685","outline-variant":"#c4c5d6","primary-fixed":"#dde1ff","error":"#ba1a1a","error-container":"#ffdad6"},borderRadius:{DEFAULT:"0.125rem",lg:"0.25rem",xl:"0.5rem",full:"0.75rem"},fontFamily:{headline:["Inter","sans-serif"],body:["Inter","sans-serif"]}}}}
</script>
<style>
.material-symbols-outlined{font-variation-settings:'FILL' 0,'wght' 400,'GRAD' 0,'opsz' 24}
.kinetic-gradient{background:linear-gradient(135deg,#1039b9 0%,#3454d1 100%)}
.ghost-border{border:2px dashed rgba(196,197,214,0.5)}
.ghost-border.drag-over{border-color:#3454d1;background:rgba(52,84,209,0.04)}
.upload-zone input[type=file]{display:none}
.status-msg{margin-top:12px;padding:10px 16px;border-radius:8px;font-size:13px;display:none}
.status-msg.success{display:block;background:#f0fdf4;color:#166534;border:1px solid #bbf7d0}
.status-msg.error{display:block;background:#fef2f2;color:#991b1b;border:1px solid #fecaca}
.status-msg.loading{display:block;background:#eff6ff;color:#1e40af;border:1px solid #bfdbfe}
</style>
</head>
<body class="bg-[#f7f9ff] font-body text-[#091d2e] selection:bg-blue-200">
<!-- Header -->
<header class="sticky top-0 w-full z-50 bg-white/70 backdrop-blur-xl shadow-sm">
<div class="flex justify-between items-center h-16 px-6 lg:px-12">
<div class="flex items-center gap-8">
<img src="/static/logo.png" alt="MELKO" class="h-10"/>
<nav class="hidden md:flex gap-6 items-center">
<a class="text-blue-700 font-bold border-b-2 border-blue-600" href="#">Dashboard</a>
<a class="text-slate-500 hover:text-blue-600 transition-colors" href="#">History</a>
<a class="text-slate-500 hover:text-blue-600 transition-colors" href="#">Settings</a>
</nav>
</div>
<div class="flex items-center gap-4">
<button class="material-symbols-outlined p-2 rounded-full hover:bg-blue-50/50 text-slate-900">help</button>
<button class="material-symbols-outlined p-2 rounded-full hover:bg-blue-50/50 text-slate-900">account_circle</button>
</div>
</div>
</header>

<div class="flex">
<!-- Sidebar -->
<aside class="h-[calc(100vh-64px)] w-64 fixed left-0 top-16 bg-slate-50 flex-col py-8 gap-4 hidden lg:flex">
<div class="px-8 mb-4">
<h2 class="text-lg font-bold text-slate-900">Packing Tool</h2>
<p class="text-[10px] uppercase font-bold tracking-widest text-slate-400">v2.1 Kinetic</p>
</div>
<nav class="flex flex-col gap-2">
<div id="nav1" class="flex items-center gap-3 bg-white text-blue-700 rounded-l-full py-3 px-4 shadow-sm ml-4 cursor-pointer transition-all">
<span class="material-symbols-outlined">upload_file</span>
<span class="text-[10px] uppercase font-bold tracking-widest">1. Generate List</span>
</div>
<div id="nav2" class="flex items-center gap-3 text-slate-500 py-3 px-4 ml-4 cursor-pointer hover:text-blue-600 transition-all">
<span class="material-symbols-outlined">fact_check</span>
<span class="text-[10px] uppercase font-bold tracking-widest">2. Review Data</span>
</div>
<div id="nav3" class="flex items-center gap-3 text-slate-500 py-3 px-4 ml-4 cursor-pointer hover:text-blue-600 transition-all">
<span class="material-symbols-outlined">description</span>
<span class="text-[10px] uppercase font-bold tracking-widest">3. Export BOL</span>
</div>
</nav>
</aside>

<!-- Main Content -->
<main class="lg:ml-64 flex-1 min-h-screen p-6 lg:p-12">
<div class="max-w-5xl mx-auto space-y-12">
<!-- Intro -->
<div class="space-y-2">
<h1 class="text-4xl lg:text-5xl font-black tracking-tight">Packing List Maker</h1>
<p class="text-[#444654] max-w-2xl leading-relaxed">上传客户的预报表/派送计划 Excel，自动生成标准 Packing List 和 BOL 文件。</p>
</div>

<div class="grid grid-cols-1 xl:grid-cols-3 gap-12">
<div class="xl:col-span-2 space-y-16">

<!-- Step 1 -->
<section class="relative pl-12">
<div class="absolute left-4 top-0 bottom-0 w-1 bg-[#dde1ff] rounded-full">
<div class="absolute top-0 left-1/2 -translate-x-1/2 w-4 h-4 rounded-full bg-[#684000] border-4 border-white"></div>
</div>
<div class="space-y-6">
<header>
<h2 class="text-2xl font-bold tracking-tight">Step 1: 生成 Packing List</h2>
<p class="text-[#444654] text-sm">上传客户的 Excel 源文件 (.xlsx / .xls)</p>
</header>
<div id="zone1" class="upload-zone group relative bg-[#edf4ff] rounded-xl p-12 ghost-border transition-all hover:border-[#3454d1] cursor-pointer" onclick="document.getElementById('file1').click()">
<input type="file" id="file1" accept=".xlsx,.xls" onchange="handleStep1(this)"/>
<div class="flex flex-col items-center text-center space-y-4">
<div class="w-16 h-16 rounded-full bg-white flex items-center justify-center shadow-sm">
<span class="material-symbols-outlined text-[#3454d1] text-3xl">upload_file</span>
</div>
<div>
<p class="font-bold">拖拽客户 Excel 文件到此处</p>
<p class="text-[#444654] text-xs">支持 .XLSX, .XLS 格式</p>
</div>
<button class="px-6 py-2 bg-white border border-[#c4c5d6] text-[#1039b9] rounded-lg text-sm font-bold hover:bg-blue-50 transition-colors" onclick="event.stopPropagation();document.getElementById('file1').click()">
浏览文件
</button>
<p class="text-xs text-[#3454d1] font-medium" id="fname1"></p>
</div>
</div>
<div class="status-msg" id="status1"></div>
</div>
</section>

<!-- Step 2 -->
<section class="relative pl-12">
<div class="absolute left-4 top-0 bottom-0 w-1 bg-[#dde1ff] rounded-full">
<div class="absolute top-0 left-1/2 -translate-x-1/2 w-4 h-4 rounded-full bg-[#dde1ff] border-4 border-white"></div>
</div>
<div class="space-y-6">
<header>
<h2 class="text-2xl font-bold tracking-tight text-slate-400">Step 2: 检查 Packing List</h2>
</header>
<div class="bg-white rounded-xl p-6 border border-slate-200/50 flex items-start gap-4">
<span class="material-symbols-outlined text-[#4e6073]">info</span>
<div>
<p class="text-sm font-semibold">打开下载的 Packing List</p>
<p class="text-sm text-[#444654]">人工检查数据是否正确，修改后保存。然后进入 Step 3 上传修改好的文件。</p>
</div>
</div>
</div>
</section>

<!-- Step 3 -->
<section class="relative pl-12">
<div class="absolute left-4 top-0 h-4 w-1 bg-[#dde1ff] rounded-t-full">
<div class="absolute top-0 left-1/2 -translate-x-1/2 w-4 h-4 rounded-full bg-[#dde1ff] border-4 border-white"></div>
</div>
<div class="space-y-6">
<header>
<h2 class="text-2xl font-bold tracking-tight text-slate-400" id="step3title">Step 3: 生成 BOL 文件</h2>
</header>
<div id="zone2" class="upload-zone group relative bg-[#edf4ff] rounded-xl p-12 ghost-border transition-all hover:border-[#3454d1] cursor-pointer" onclick="document.getElementById('file2').click()">
<input type="file" id="file2" accept=".xlsx" onchange="handleStep2(this)"/>
<div class="flex flex-col items-center text-center space-y-4">
<div class="w-16 h-16 rounded-full bg-white flex items-center justify-center shadow-sm">
<span class="material-symbols-outlined text-[#3454d1] text-3xl">inventory_2</span>
</div>
<div>
<p class="font-bold">拖拽修改好的 Packing List 到此处</p>
<p class="text-[#444654] text-xs">上传后自动生成 BOL 文件包 (ZIP 下载)</p>
</div>
<button class="px-6 py-2 bg-white border border-[#c4c5d6] text-[#1039b9] rounded-lg text-sm font-bold hover:bg-blue-50 transition-colors" onclick="event.stopPropagation();document.getElementById('file2').click()">
浏览文件
</button>
<p class="text-xs text-[#3454d1] font-medium" id="fname2"></p>
</div>
</div>
<div class="status-msg" id="status2"></div>
</div>
</section>

</div>

<!-- Sidebar Instructions -->
<aside class="space-y-8">
<div class="bg-[#3454d1] text-white p-8 rounded-xl shadow-xl shadow-blue-500/10">
<h3 class="text-lg font-bold mb-4 flex items-center gap-2">
<span class="material-symbols-outlined">auto_awesome</span>
自动化功能
</h3>
<p class="text-sm leading-relaxed opacity-80 mb-6">系统自动识别客户 Excel 格式，转换重量单位，按派送方式分类着色。</p>
<div class="space-y-4">
<div class="flex gap-3">
<span class="material-symbols-outlined text-amber-300">check_circle</span>
<div class="text-xs"><p class="font-bold">KG → LBS 自动转换</p><p class="opacity-70">×2.2 自动计算 CTN/LBS 和总 LBS</p></div>
</div>
<div class="flex gap-3">
<span class="material-symbols-outlined text-amber-300">check_circle</span>
<div class="text-xs"><p class="font-bold">颜色自动标记</p><p class="opacity-70">FEDEX紫/UPS褐/HOLD红/FF黄</p></div>
</div>
<div class="flex gap-3">
<span class="material-symbols-outlined text-amber-300">check_circle</span>
<div class="text-xs"><p class="font-bold">仓库代码自动合并</p><p class="opacity-70">FBA CODE 连续相同值合并单元格</p></div>
</div>
<div class="flex gap-3">
<span class="material-symbols-outlined text-amber-300">check_circle</span>
<div class="text-xs"><p class="font-bold">支持 .xls + .xlsx</p><p class="opacity-70">兼容多种客户模板格式</p></div>
</div>
</div>
</div>
<div class="bg-white p-8 rounded-xl border border-slate-200/50 space-y-4">
<h3 class="text-sm font-bold uppercase tracking-widest text-[#4e6073]">BOL 文件包含</h3>
<ul class="space-y-4">
<li class="flex items-start gap-3">
<span class="w-5 h-5 rounded bg-slate-100 flex items-center justify-center text-[10px] font-bold text-slate-500 shrink-0 mt-0.5">01</span>
<p class="text-xs text-[#444654]"><strong>Packing List</strong> (A-M 13列精简版)</p>
</li>
<li class="flex items-start gap-3">
<span class="w-5 h-5 rounded bg-slate-100 flex items-center justify-center text-[10px] font-bold text-slate-500 shrink-0 mt-0.5">02</span>
<p class="text-xs text-[#444654]"><strong>Hold List</strong> (仅 HOLD/RELABEL 行)</p>
</li>
<li class="flex items-start gap-3">
<span class="w-5 h-5 rounded bg-slate-100 flex items-center justify-center text-[10px] font-bold text-slate-500 shrink-0 mt-0.5">03</span>
<p class="text-xs text-[#444654]"><strong>TRUCKING.docx</strong> (仓库板数/CBM 汇总)</p>
</li>
<li class="flex items-start gap-3">
<span class="w-5 h-5 rounded bg-slate-100 flex items-center justify-center text-[10px] font-bold text-slate-500 shrink-0 mt-0.5">04</span>
<p class="text-xs text-[#444654]"><strong>柜号唛头.docx</strong> (柜号 + MELKO)</p>
</li>
</ul>
</div>
</aside>
</div>
</div>
</main>
</div>

<!-- Footer -->
<footer class="w-full py-6 mt-auto bg-slate-50 border-t border-slate-200/50">
<div class="flex flex-col md:flex-row justify-between items-center px-12 gap-4">
<p class="text-[10px] uppercase font-bold tracking-widest text-slate-400">&copy; 2026 MELKO Logistics</p>
</div>
</footer>

<script>
// Drag-drop for both zones
['zone1','zone2'].forEach(id=>{
const zone=document.getElementById(id);
zone.addEventListener('dragover',e=>{e.preventDefault();zone.classList.add('drag-over')});
zone.addEventListener('dragleave',()=>zone.classList.remove('drag-over'));
zone.addEventListener('drop',e=>{
e.preventDefault();zone.classList.remove('drag-over');
const file=e.dataTransfer.files[0];if(!file)return;
const input=zone.querySelector('input[type=file]');
const dt=new DataTransfer();dt.items.add(file);input.files=dt.files;
input.dispatchEvent(new Event('change'));
});
});

function setStatus(id,type,msg){
const el=document.getElementById(id);
el.className='status-msg '+type;el.textContent=msg;
}

async function handleStep1(input){
const file=input.files[0];if(!file)return;
document.getElementById('fname1').textContent=file.name;
setStatus('status1','loading','正在处理...');
const form=new FormData();form.append('file',file);
try{
const resp=await fetch('/api/step1',{method:'POST',body:form});
if(!resp.ok){const err=await resp.json();throw new Error(err.error||'Unknown error')}
const blob=await resp.blob();
const cd=resp.headers.get('Content-Disposition')||'';
const match=cd.match(/filename="?([^"]+)"?/);
const fname=match?match[1]:'packing_list.xlsx';
const url=URL.createObjectURL(blob);
const a=document.createElement('a');a.href=url;a.download=decodeURIComponent(fname);a.click();
URL.revokeObjectURL(url);
setStatus('status1','success','✅ 已生成: '+decodeURIComponent(fname)+' — 已自动下载，请检查后进入 Step 3');
}catch(e){setStatus('status1','error','❌ '+e.message)}
input.value='';
}

async function handleStep2(input){
const file=input.files[0];if(!file)return;
document.getElementById('fname2').textContent=file.name;
setStatus('status2','loading','正在生成 BOL 文件包...');
const form=new FormData();form.append('file',file);
try{
const resp=await fetch('/api/step2',{method:'POST',body:form});
if(!resp.ok){const err=await resp.json();throw new Error(err.error||'Unknown error')}
const blob=await resp.blob();
const cd=resp.headers.get('Content-Disposition')||'';
const match=cd.match(/filename="?([^"]+)"?/);
const fname=match?match[1]:'BOL.zip';
const url=URL.createObjectURL(blob);
const a=document.createElement('a');a.href=url;a.download=decodeURIComponent(fname);a.click();
URL.revokeObjectURL(url);
setStatus('status2','success','✅ BOL 文件包已下载: '+decodeURIComponent(fname));
}catch(e){setStatus('status2','error','❌ '+e.message)}
input.value='';
}
</script>
</body></html>
"""


@app.route('/')
def index():
    return render_template_string(HTML)


@app.route('/api/step1', methods=['POST'])
def step1():
    """Upload source Excel → return Packing List xlsx."""
    if 'file' not in request.files:
        return jsonify(error="No file uploaded"), 400

    f = request.files['file']
    if not f.filename:
        return jsonify(error="Empty filename"), 400

    ext = os.path.splitext(f.filename)[1].lower()
    if ext not in ('.xlsx', '.xls'):
        return jsonify(error=f"Unsupported file type: {ext}"), 400

    tmpdir = tempfile.mkdtemp()
    try:
        src_path = os.path.join(tmpdir, f.filename)
        f.save(src_path)

        # Read and transform
        raw_headers, rows, header_map, header_to_canon, metadata = converter.read_source(src_path)
        container_no = metadata.get("container") or converter._extract_container(f.filename)
        sorted_rows = converter.sort_rows(rows, header_map)
        transformed = [converter.transform_row(r, header_map) for r in sorted_rows]

        all_candidates = [
            h for h in raw_headers
            if header_to_canon.get(h) is None and h
            and converter._norm_header_key(h) not in converter.SKIP_HEADER_KEYS
        ]
        extras_headers = [
            h for h in all_candidates
            if any(sr.get(h) not in (None, "") for sr in sorted_rows)
        ]

        import openpyxl
        out_path = os.path.join(tmpdir, f"{container_no} packing list.xlsx")
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "\u603b\u8868"
        for col_letter, w in converter.COLUMN_WIDTHS.items():
            ws.column_dimensions[col_letter].width = w
        converter._write_header_block(ws, container_no, metadata, extras_headers)
        last_data = converter._write_data_rows(ws, transformed, sorted_rows, extras_headers)
        converter._write_total_row(ws, last_data, transformed, len(extras_headers))
        ws.freeze_panes = "A4"
        wb.save(out_path)

        return send_file(
            out_path,
            as_attachment=True,
            download_name=f"{container_no} packing list.xlsx",
            mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
        )
    except Exception as e:
        return jsonify(error=str(e)), 500
    finally:
        shutil.rmtree(tmpdir, ignore_errors=True)


@app.route('/api/step2', methods=['POST'])
def step2():
    """Upload reviewed Packing List → return ZIP of BOL files."""
    if 'file' not in request.files:
        return jsonify(error="No file uploaded"), 400

    f = request.files['file']
    if not f.filename:
        return jsonify(error="Empty filename"), 400

    tmpdir = tempfile.mkdtemp()
    try:
        src_path = os.path.join(tmpdir, f.filename)
        f.save(src_path)

        import openpyxl
        wb = openpyxl.load_workbook(src_path, data_only=True)
        ws = wb.active

        container_no = converter._norm_str(ws.cell(row=1, column=1).value)
        if not container_no:
            container_no = converter._extract_container(f.filename)

        metadata = {"container": container_no, "etd": "", "eta": ""}
        etd_val = ws.cell(row=2, column=6).value
        eta_val = ws.cell(row=2, column=11).value
        if etd_val:
            metadata["etd"] = converter._norm_str(etd_val)
        if eta_val:
            metadata["eta"] = converter._norm_str(eta_val)

        # Read data
        transformed = []
        for row in range(4, ws.max_row + 1):
            method = converter._norm_str(ws.cell(row=row, column=3).value)
            if method == "TOTAL" or not method:
                continue
            ship_id = converter._norm_str(ws.cell(row=row, column=2).value)
            if not ship_id:
                continue
            r = {
                "no": converter._norm_str(ws.cell(row=row, column=1).value),
                "ship_id": ship_id,
                "method": method,
                "ctns": converter._to_num(ws.cell(row=row, column=4).value),
                "cbm": converter._to_num(ws.cell(row=row, column=5).value),
                "kg": converter._to_num(ws.cell(row=row, column=6).value),
                "ctn_lbs": converter._to_num(ws.cell(row=row, column=7).value),
                "total_lbs": converter._to_num(ws.cell(row=row, column=8).value),
                "description": converter._norm_str(ws.cell(row=row, column=9).value),
                "fba_code": converter._norm_str(ws.cell(row=row, column=10).value),
                "address": converter._norm_str(ws.cell(row=row, column=11).value),
                "fba_id": converter._norm_str(ws.cell(row=row, column=12).value),
                "reference_id": converter._norm_str(ws.cell(row=row, column=13).value),
                "loading_order": converter._norm_str(ws.cell(row=row, column=14).value),
                "eta_date": converter._norm_str(ws.cell(row=row, column=15).value),
                "trucking_no": converter._norm_str(ws.cell(row=row, column=16).value),
                "rate": converter._norm_str(ws.cell(row=row, column=17).value),
                "remark": converter._norm_str(ws.cell(row=row, column=18).value),
            }
            if r["ctns"] == int(r["ctns"]):
                r["ctns"] = int(r["ctns"])
            transformed.append(r)

        # Generate BOL files
        bol_dir = os.path.join(tmpdir, "BOL")
        os.makedirs(bol_dir)

        files_generated = []

        bol_pl = os.path.join(bol_dir, f"{container_no} packing list.xlsx")
        converter.build_bol_packing_list(transformed, container_no, bol_pl)
        files_generated.append(bol_pl)

        bol_hl = os.path.join(bol_dir, f"{container_no} hold list.xlsx")
        if converter.build_bol_hold_list(transformed, container_no, bol_hl):
            files_generated.append(bol_hl)

        bol_tr = os.path.join(bol_dir, f"{container_no}-TRUCKING.docx")
        converter.build_bol_trucking(transformed, container_no, bol_tr)
        files_generated.append(bol_tr)

        bol_sm = os.path.join(bol_dir, f"{container_no}-\u67dc\u53f7\u551b\u5934.docx")
        converter.build_bol_ship_marks(container_no, bol_sm)
        files_generated.append(bol_sm)

        # Create ZIP
        zip_buffer = io.BytesIO()
        with zipfile.ZipFile(zip_buffer, 'w', zipfile.ZIP_DEFLATED) as zf:
            for fp in files_generated:
                zf.write(fp, os.path.basename(fp))
        zip_buffer.seek(0)

        return send_file(
            zip_buffer,
            as_attachment=True,
            download_name=f"{container_no} BOL.zip",
            mimetype='application/zip',
        )
    except Exception as e:
        return jsonify(error=str(e)), 500
    finally:
        shutil.rmtree(tmpdir, ignore_errors=True)


if __name__ == '__main__':
    port = int(os.environ.get('PORT', 5000))
    app.run(host='0.0.0.0', port=port, debug=False)
